"""xlsx_report — typed Excel report with a native, auto-selected chart.

Reverse-engineered from the plugin's ``visualizer.ts``. That file hand-built
OOXML chart XML and stitched it into the SheetJS ZIP with ``adm-zip`` ONLY
because SheetJS cannot embed charts. openpyxl has native chart support, so the
entire OOXML-injection machinery collapses to ``openpyxl.chart.LineChart /
BarChart`` + typed cells. We port the CONTRACT, not the mechanism:

  * typed cells      — dates as real ``datetime`` values carrying a date
                       number-format (``cell.is_date`` True); ints/reals as
                       numbers; headers as strings
  * sheet chrome     — auto-filter + frozen header row on a "QueryResults" sheet
  * chart selection  — datetime+numeric -> LineChart (<=3 series);
                       categorical+numeric -> BarChart, >=3 numeric cols ->
                       column variant, else bar variant; else no chart
  * empty-rowset guard — never add a chart on 0 rows (the TS ``lastRow`` guard:
                       an empty range like ``$B$2:$B$1`` is invalid)
  * ExecutingUser normalization in the export column
  * NON-FATAL chart failure — if chart building raises, the data sheet is still
                       valid and the workbook bytes are returned

Like the plugin's desktop file-write, but hosted: ``build_xlsx_report`` RETURNS
the ``.xlsx`` bytes; the caller decides where they go. ``openpyxl`` is a lazy
import so importing this module never hard-requires it; a missing install
raises a clear ``ImportError`` telling the caller to add openpyxl.
"""
from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Optional, Sequence

__all__ = ["build_xlsx_report", "normalize_executing_user_display"]

_DATE_NUMBER_FORMAT = "yyyy-mm-dd hh:mm:ss"

# ── ExecutingUser normalization (mirrors constants.ts / format.ts) ──────────
EXECUTING_USER_COLUMN = "ExecutingUser"
from ..identity_display import (  # noqa: F401  (re-exported: see identity_display)
    NEWELL_EMAIL_DOMAIN, normalize_executing_user_display)
_EXECUTING_USER_LOWER = EXECUTING_USER_COLUMN.lower()


# ── Lazy openpyxl import ────────────────────────────────────────────────────
def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:  # pragma: no cover - environment-dependent
        raise ImportError(
            "build_xlsx_report requires the 'openpyxl' package, which is not "
            "installed. Add it to the environment (e.g. `pip install openpyxl`)."
        ) from exc
    return openpyxl


# ── Column shape helpers ────────────────────────────────────────────────────
def _col_name(col: Any) -> str:
    name = col.get("name") if isinstance(col, dict) else getattr(col, "name", None)
    return "" if name is None else str(name)


def _col_type(col: Any) -> str:
    ctype = col.get("type") if isinstance(col, dict) else getattr(col, "type", None)
    return "" if ctype is None else str(ctype)


def _is_datetime(col: Any) -> bool:
    return re.search(r"datetime", _col_type(col), re.IGNORECASE) is not None


def _is_numeric(col: Any) -> bool:
    return re.match(r"(int|long|real|double|decimal)\b", _col_type(col), re.IGNORECASE) is not None


# ── Classification (visualizer.ts contract) ─────────────────────────────────
@dataclass
class _ColumnClassification:
    visual_type: str  # "lineChart" | "clusteredColumnChart" | "clusteredBarChart" | "tableEx"
    category_idx: int
    value_indices: list


def _classify_columns(columns: Sequence[Any]) -> _ColumnClassification:
    datetime_idx = [i for i, c in enumerate(columns) if _is_datetime(c)]
    numeric_idx = [i for i, c in enumerate(columns) if _is_numeric(c)]
    categorical_idx = [
        i for i, c in enumerate(columns) if not _is_datetime(c) and not _is_numeric(c)
    ]

    if datetime_idx and numeric_idx:
        return _ColumnClassification("lineChart", datetime_idx[0], list(numeric_idx[:3]))

    if categorical_idx and numeric_idx:
        v_type = "clusteredColumnChart" if len(numeric_idx) >= 3 else "clusteredBarChart"
        return _ColumnClassification(v_type, categorical_idx[0], list(numeric_idx[:4]))

    return _ColumnClassification("tableEx", -1, [])


# ── Cell coercion (port of visualizer.ts coerceCell) ────────────────────────
def _coerce_datetime(raw: Any) -> Any:
    # Excel has no timezone concept — datetimes stored in a worksheet MUST be
    # naive (openpyxl raises on tz-aware values). Strip tzinfo, keeping the
    # parsed wall-clock, exactly as SheetJS wrote the plugin's Date values.
    if isinstance(raw, datetime):
        return raw.replace(tzinfo=None)
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    candidate = s[:-1] + "+00:00" if s.endswith("Z") else s
    try:
        parsed = datetime.fromisoformat(candidate)
    except ValueError:
        return str(raw)  # unparseable -> keep as string (TS: NaN date -> String(raw))
    return parsed.replace(tzinfo=None)


def _coerce_number(raw: Any) -> Any:
    if isinstance(raw, bool):
        return str(raw)
    if isinstance(raw, (int, float)):
        return raw
    s = str(raw)
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        return str(raw)  # NaN -> String(raw)


def _coerce_cell(raw: Any, col_type: str) -> Any:
    if raw is None:
        return None
    t = col_type.lower()
    if re.search(r"datetime", t):
        return _coerce_datetime(raw)
    if re.match(r"(int|long)\b", t):
        return _coerce_number(raw)
    if re.match(r"(real|double|decimal)\b", t):
        return _coerce_number(raw)
    if t.startswith("bool"):
        if isinstance(raw, bool):
            return raw
        s = str(raw).lower()
        if s == "true":
            return True
        if s == "false":
            return False
        return str(raw)
    if t.startswith("dynamic"):
        if isinstance(raw, (dict, list)):
            return json.dumps(raw, separators=(",", ":"))
        return str(raw)
    return str(raw)


# ── Chart building (module-level so tests can monkeypatch it to raise) ──────
def _build_and_add_chart(
    ws,
    cls: _ColumnClassification,
    columns: Sequence[Any],
    row_count: int,
    title: str,
) -> None:
    """Add a native openpyxl chart to ``ws`` per the classification.

    Isolated as a module-level function so the caller can wrap it in try/except
    (non-fatal chart failure) and so tests can monkeypatch it to raise while
    still expecting a valid workbook.
    """
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter

    last_row = row_count + 1  # +1 for the header row

    if cls.visual_type == "lineChart":
        chart = LineChart()
    else:
        chart = BarChart()
        # clusteredColumnChart -> vertical columns; clusteredBarChart -> horizontal bars
        chart.type = "col" if cls.visual_type == "clusteredColumnChart" else "bar"
        chart.grouping = "clustered"

    chart.title = title
    chart.style = 10

    # One data reference per value column (columns need not be contiguous).
    for value_idx in cls.value_indices:
        col_1based = value_idx + 1
        ref = Reference(ws, min_col=col_1based, max_col=col_1based, min_row=1, max_row=last_row)
        chart.add_data(ref, titles_from_data=True)

    cat_col_1based = cls.category_idx + 1
    cats = Reference(ws, min_col=cat_col_1based, max_col=cat_col_1based, min_row=2, max_row=last_row)
    chart.set_categories(cats)

    # Anchor to the right of the data area (mirrors the TS drawing anchor intent).
    anchor_col = max(9, len(columns) + 2)
    ws.add_chart(chart, f"{get_column_letter(anchor_col)}2")


# ── Public builder ──────────────────────────────────────────────────────────
def build_xlsx_report(
    columns: Sequence[Any],
    rows: Sequence[Sequence[Any]],
    title: str = "Query Results",
) -> bytes:
    """Build a typed ``.xlsx`` workbook with an auto-selected chart; RETURN bytes.

    Performs no disk I/O (unlike the plugin's file-write). The data sheet is
    always valid even if chart building fails (non-fatal).
    """
    openpyxl = _require_openpyxl()
    columns = list(columns)
    rows = list(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QueryResults"

    # Header row — always strings.
    ws.append([_col_name(c) for c in columns])

    exec_user_cols = {i for i, c in enumerate(columns) if _col_name(c).lower() == _EXECUTING_USER_LOWER}

    # Data rows — typed per column.
    for row in rows:
        out_row = []
        for ci, col in enumerate(columns):
            raw = row[ci] if ci < len(row) else None
            if ci in exec_user_cols:
                out_row.append(None if raw is None else normalize_executing_user_display(raw))
            else:
                out_row.append(_coerce_cell(raw, _col_type(col)))
        ws.append(out_row)

    # Date number-format on datetime columns so Excel reads them as DateTime
    # (the inspect gate checks numeric-stored + numFmt; openpyxl produces this
    # natively when a datetime value carries a date number_format).
    datetime_col_idx = [i for i, c in enumerate(columns) if _is_datetime(c)]
    for ci in datetime_col_idx:
        for ri in range(2, len(rows) + 2):  # data rows start at Excel row 2
            cell = ws.cell(row=ri, column=ci + 1)
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = _DATE_NUMBER_FORMAT

    # Column widths (cosmetic parity with the TS !cols).
    from openpyxl.utils import get_column_letter

    for ci, col in enumerate(columns):
        ws.column_dimensions[get_column_letter(ci + 1)].width = max(len(_col_name(col)) + 2, 16)

    # Auto-filter over the whole used range + frozen header row.
    if columns:
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Chart — only when the shape supports it AND there is at least one row.
    # Failure is non-fatal: the data sheet must still be valid and returned.
    cls = _classify_columns(columns)
    if cls.visual_type != "tableEx" and cls.value_indices and len(rows) > 0:
        try:
            _build_and_add_chart(ws, cls, columns, len(rows), title)
        except Exception:
            # Swallow — the typed data sheet is already complete and valid.
            pass

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
