"""artifact_lookup.py — three-way lookup over ArtifactsMappedtoWorkspace.xlsx.

Python port of the plugin's ``services/artifact-lookup.ts``. Loads and queries the
artifact inventory, providing three lookup directions (by name, id, or workspace) for
the artifact-lookup tool.

Naming: ``pbiWorkspaceName`` is the Power BI report workspace — NOT the Azure Log
Analytics workspace the collector connects to.

Expected Excel columns (case-insensitive match at load time):
    ArtifactId           — unique identifier for the dataset
    ArtifactName         — display name (matches canonicalName in routing_table.py)
    PowerBIWorkspaceName — the Power BI workspace the artifact lives in

Dedup rule: if the same ArtifactId appears in multiple rows with different
PowerBIWorkspaceNames, the FIRST-encountered row wins and a conflict message is logged to
stderr. Does not fail startup.

Shared normalisation: ``normalize_for_matching`` is used for artifact_name and
pbi_workspace_name lookups (case-insensitive, punctuation-agnostic). artifact_id uses
exact trimmed-string matching.

xlsx reading: openpyxl is imported LAZILY inside ``load()``. If openpyxl is not
installed, the loader degrades gracefully — the store stays unloaded and every lookup
returns ``unavailable`` — rather than raising at import time. The dedup / lookup logic
is independently testable via ``load_from_rows``.
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Dict, List, Optional

from . import artifacts_xlsx_path as _default_xlsx_path
from .text_normalize import normalize_for_matching


class ArtifactLookup:
    """Three-way artifact inventory lookup with first-encountered dedup.

    Load from an xlsx path (openpyxl, lazy) or directly from a list of row dicts (for
    tests / injection). Degrades to ``unavailable`` on any load failure.
    """

    def __init__(self, xlsx_path: Optional[Path] = None, *, load: bool = False) -> None:
        self._xlsx_path: Path = Path(xlsx_path) if xlsx_path is not None else _default_xlsx_path()
        # None = not yet loaded (or failed to load).
        self._store: Optional[List[Dict]] = None
        self._load_error: Optional[str] = None
        if load:
            self.load()

    # ── Loaders ────────────────────────────────────────────────────────────────
    def load(self, xlsx_path: Optional[Path] = None) -> None:
        """Load the inventory from the xlsx via openpyxl (imported lazily).

        Non-fatal on any failure (missing file, missing openpyxl, bad shape): sets
        ``load_error`` and leaves the store unloaded so lookups report ``unavailable``.
        """
        path = Path(xlsx_path) if xlsx_path is not None else self._xlsx_path
        try:
            try:
                from openpyxl import load_workbook  # lazy import — optional dependency
            except ImportError:
                self._load_error = (
                    "openpyxl is not installed — artifact lookup unavailable. "
                    "Install the [prod] extra to enable xlsx reading."
                )
                sys.stderr.write(f"ArtifactLookup load error: {self._load_error}\n")
                return

            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]] if wb.sheetnames else None
            if ws is None:
                self._load_error = "ArtifactsMappedtoWorkspace.xlsx has no sheets — cannot load artifact data."
                return

            row_iter = ws.iter_rows(values_only=True)
            try:
                header = next(row_iter)
            except StopIteration:
                self._store = []
                wb.close()
                return
            header_keys = [str(h).strip() if h is not None else "" for h in header]
            rows: List[Dict] = []
            for values in row_iter:
                row: Dict = {}
                for i, key in enumerate(header_keys):
                    if key:
                        row[key] = values[i] if i < len(values) else None
                rows.append(row)
            wb.close()
            self.load_from_rows(rows)
        except Exception as err:
            self._load_error = str(err)
            sys.stderr.write(f"ArtifactLookup load error: {self._load_error}\n")

    def load_from_rows(self, raw_rows: List[Dict]) -> None:
        """Build the store from row dicts (header-keyed), applying dedup + conflict logging.

        Mirrors the TS ``loadArtifacts`` core: case-insensitive column resolution,
        first-encountered dedup by ArtifactId, one stderr line per conflict. Directly
        testable without openpyxl.
        """
        first_row = raw_rows[0] if raw_rows else {}
        keys = list(first_row.keys())

        def find_col(expected: str) -> Optional[str]:
            lower = expected.lower()
            return next((k for k in keys if k.lower() == lower), None)

        col_id = find_col("ArtifactId")
        col_name = find_col("ArtifactName")
        col_workspace = find_col("PowerBIWorkspaceName")

        if not col_id or not col_name or not col_workspace:
            missing = ", ".join(
                name
                for present, name in (
                    (col_id, "ArtifactId"),
                    (col_name, "ArtifactName"),
                    (col_workspace, "PowerBIWorkspaceName"),
                )
                if not present
            )
            self._load_error = (
                f"ArtifactsMappedtoWorkspace.xlsx is missing required columns: {missing}. "
                f"Found: {', '.join(keys)}."
            )
            return

        seen: Dict[str, Dict] = {}
        conflicts: Dict[str, List[str]] = {}

        for row in raw_rows:
            artifact_id = str(row.get(col_id) if row.get(col_id) is not None else "").strip()
            artifact_name = str(row.get(col_name) if row.get(col_name) is not None else "").strip()
            pbi_workspace_name = str(row.get(col_workspace) if row.get(col_workspace) is not None else "").strip()

            if not artifact_id:
                continue  # skip rows with no ArtifactId

            existing = seen.get(artifact_id)
            if existing is not None:
                # Record the conflict; list starts with the first-encountered workspace.
                lst = conflicts.get(artifact_id)
                if lst is not None:
                    lst.append(pbi_workspace_name)
                else:
                    conflicts[artifact_id] = [existing["pbiWorkspaceName"], pbi_workspace_name]
            else:
                seen[artifact_id] = {
                    "artifactId": artifact_id,
                    "artifactName": artifact_name,
                    "pbiWorkspaceName": pbi_workspace_name,
                }

        for artifact_id, workspaces in conflicts.items():
            first = workspaces[0] if workspaces else "(unknown)"
            sys.stderr.write(
                f"ArtifactDataset conflict: ArtifactId {artifact_id} appears under multiple "
                f"PowerBIWorkspaceNames: {', '.join(workspaces)}. Using first-encountered ({first}) "
                "for lookups. Verify the source data.\n"
            )

        self._store = list(seen.values())
        self._load_error = None

    def is_available(self) -> bool:
        return self._store is not None

    @property
    def load_error(self) -> Optional[str]:
        return self._load_error

    # ── Lookup ────────────────────────────────────────────────────────────────
    def lookup(
        self,
        artifact_name: Optional[str] = None,
        artifact_id: Optional[str] = None,
        pbi_workspace_name: Optional[str] = None,
    ) -> Dict:
        """Three-way lookup. Exactly ONE parameter must be supplied (else raises).

        Statuses: ``found`` | ``found_workspace`` | ``multiple`` | ``multiple_workspaces``
        | ``not_found`` | ``unavailable``.
        """
        provided = [v for v in (artifact_name, artifact_id, pbi_workspace_name) if v is not None]
        if len(provided) != 1:
            raise ValueError(
                "lookup requires exactly one of artifact_name, artifact_id, or "
                f"pbi_workspace_name (got {len(provided)})."
            )

        if self._store is None:
            reason = self._load_error or "Artifact data not loaded — load() was not called at startup."
            return {"status": "unavailable", "reason": reason, "message": f"Artifact lookup unavailable: {reason}"}

        # ── artifact_id: exact trimmed match ──
        if artifact_id is not None:
            id_ = artifact_id.strip()
            match = next((a for a in self._store if a["artifactId"] == id_), None)
            if match is None:
                return {"status": "not_found", "query": id_, "message": f"No artifact found with ArtifactId '{id_}'."}
            return {
                "status": "found",
                "artifact": match,
                "message": (
                    f"Found: {match['artifactName']} (ArtifactId: {match['artifactId']}) "
                    f"in the {match['pbiWorkspaceName']} workspace."
                ),
            }

        # ── artifact_name: normalized exact match ──
        if artifact_name is not None:
            query = artifact_name.strip()
            normalized = normalize_for_matching(query)
            matches = [a for a in self._store if normalize_for_matching(a["artifactName"]) == normalized]
            if not matches:
                return {"status": "not_found", "query": query, "message": f"No artifact found with name '{query}'."}
            if len(matches) == 1:
                m = matches[0]
                return {
                    "status": "found",
                    "artifact": m,
                    "message": (
                        f"Found: {m['artifactName']} (ArtifactId: {m['artifactId']}) "
                        f"in the {m['pbiWorkspaceName']} workspace."
                    ),
                }
            matches_desc = ", ".join(f"{m['artifactName']} ({m['pbiWorkspaceName']})" for m in matches)
            return {
                "status": "multiple",
                "query": query,
                "matches": matches,
                "message": (
                    f"'{query}' matched {len(matches)} artifacts: {matches_desc}. "
                    "Provide an artifact_id for an unambiguous lookup."
                ),
            }

        # ── pbi_workspace_name: normalized exact match on workspace name ──
        query = pbi_workspace_name.strip()  # type: ignore[union-attr]
        normalized = normalize_for_matching(query)
        matches = [a for a in self._store if normalize_for_matching(a["pbiWorkspaceName"]) == normalized]
        if not matches:
            return {"status": "not_found", "query": query, "message": f"No artifacts found in workspace '{query}'."}

        distinct_workspaces = list(dict.fromkeys(m["pbiWorkspaceName"] for m in matches))
        if len(distinct_workspaces) > 1:
            return {
                "status": "multiple_workspaces",
                "query": query,
                "workspaceNames": distinct_workspaces,
                "message": (
                    f"'{query}' matched multiple distinct workspace names: {', '.join(distinct_workspaces)}. "
                    "Use the exact workspace name for an unambiguous lookup."
                ),
            }

        workspace_name = distinct_workspaces[0] if distinct_workspaces else query
        plural = "s" if len(matches) != 1 else ""
        artifacts_desc = ", ".join(f"{m['artifactName']} ({m['artifactId']})" for m in matches)
        return {
            "status": "found_workspace",
            "workspaceName": workspace_name,
            "artifacts": matches,
            "message": (
                f"Found {len(matches)} artifact{plural} in the {workspace_name} workspace: {artifacts_desc}."
            ),
        }


_DEFAULT_ARTIFACT_LOOKUP: Optional[ArtifactLookup] = None


def default_artifact_lookup() -> ArtifactLookup:
    """Return a cached ``ArtifactLookup`` loaded from the default xlsx path.

    Degrades to ``unavailable`` if openpyxl is not installed or the file is missing.
    """
    global _DEFAULT_ARTIFACT_LOOKUP
    if _DEFAULT_ARTIFACT_LOOKUP is None:
        _DEFAULT_ARTIFACT_LOOKUP = ArtifactLookup(load=True)
    return _DEFAULT_ARTIFACT_LOOKUP
