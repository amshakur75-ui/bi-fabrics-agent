"""Tests for the xlsx export builder (Phase 5.3).

Covers: non-empty bytes re-openable by openpyxl, real datetime storage (is_date
/ date number_format), numeric typing, string headers, auto-filter + frozen
header row, ExecutingUser normalization in the export column, chart
auto-selection, empty-rowset guard, and NON-FATAL chart failure (a raising
chart step still yields a valid workbook).
"""
import io

import pytest

pytest.importorskip("openpyxl")

import openpyxl  # noqa: E402

from fabric_audit_agent.export import xlsx_report  # noqa: E402
from fabric_audit_agent.export.xlsx_report import (  # noqa: E402
    build_xlsx_report,
    normalize_executing_user_display,
)


def _reopen(data: bytes):
    assert isinstance(data, bytes) and len(data) > 0
    return openpyxl.load_workbook(io.BytesIO(data))


# ── Basic workbook shape ─────────────────────────────────────────────────────
def test_returns_reopenable_bytes_with_queryresults_sheet():
    cols = [{"name": "Category", "type": "string"}, {"name": "Value", "type": "long"}]
    rows = [["A", 1], ["B", 2]]
    wb = _reopen(build_xlsx_report(cols, rows))
    assert "QueryResults" in wb.sheetnames
    ws = wb["QueryResults"]
    assert [c.value for c in ws[1]] == ["Category", "Value"]  # headers as strings
    assert all(isinstance(c.value, str) for c in ws[1])


def test_header_row_frozen_and_autofilter_set():
    cols = [{"name": "Category", "type": "string"}, {"name": "Value", "type": "long"}]
    rows = [["A", 1], ["B", 2]]
    ws = _reopen(build_xlsx_report(cols, rows))["QueryResults"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


# ── Typed cells ──────────────────────────────────────────────────────────────
def test_dates_stored_as_real_datetimes_with_numfmt():
    cols = [{"name": "TimeGenerated", "type": "datetime"}, {"name": "CpuTimeMs", "type": "long"}]
    rows = [["2026-08-01T00:00:00Z", 10], ["2026-08-01T01:00:00Z", 20]]
    ws = _reopen(build_xlsx_report(cols, rows))["QueryResults"]
    date_cell = ws.cell(row=2, column=1)
    assert date_cell.is_date is True
    # a date/time number format is applied (not "General")
    assert date_cell.number_format not in ("General", None)
    num_cell = ws.cell(row=2, column=2)
    assert isinstance(num_cell.value, (int, float)) and not isinstance(num_cell.value, bool)


def test_numeric_strings_become_numbers():
    cols = [{"name": "Cat", "type": "string"}, {"name": "N", "type": "long"}, {"name": "R", "type": "real"}]
    rows = [["a", "42", "3.5"]]
    ws = _reopen(build_xlsx_report(cols, rows))["QueryResults"]
    assert ws.cell(row=2, column=2).value == 42
    assert ws.cell(row=2, column=3).value == 3.5


def test_executing_user_normalized_in_export_column():
    cols = [{"name": "ExecutingUser", "type": "string"}, {"name": "N", "type": "long"}]
    rows = [["jdoe", 1], ["a@b.com", 2], [None, 3]]
    ws = _reopen(build_xlsx_report(cols, rows))["QueryResults"]
    assert ws.cell(row=2, column=1).value == "jdoe@newellco.com"
    assert ws.cell(row=3, column=1).value == "a@b.com"
    assert ws.cell(row=4, column=1).value in (None, "")


# ── Chart auto-selection ─────────────────────────────────────────────────────
def test_line_chart_for_datetime_plus_numeric():
    cols = [{"name": "TimeGenerated", "type": "datetime"}, {"name": "V", "type": "long"}]
    rows = [["2026-08-01T00:00:00Z", 10], ["2026-08-01T01:00:00Z", 20]]
    ws = _reopen(build_xlsx_report(cols, rows))["QueryResults"]
    charts = ws._charts
    assert len(charts) == 1
    assert charts[0].__class__.__name__ == "LineChart"


def test_bar_chart_for_categorical_plus_numeric():
    cols = [{"name": "Cat", "type": "string"}, {"name": "V", "type": "long"}]
    rows = [["a", 1], ["b", 2]]
    ws = _reopen(build_xlsx_report(cols, rows))["QueryResults"]
    charts = ws._charts
    assert len(charts) == 1
    assert charts[0].__class__.__name__ == "BarChart"
    assert charts[0].type == "bar"  # <3 numeric cols -> horizontal bar variant


def test_column_variant_for_three_or_more_numeric():
    cols = [
        {"name": "Cat", "type": "string"},
        {"name": "A", "type": "long"},
        {"name": "B", "type": "long"},
        {"name": "C", "type": "real"},
    ]
    rows = [["x", 1, 2, 3.0], ["y", 4, 5, 6.0]]
    ws = _reopen(build_xlsx_report(cols, rows))["QueryResults"]
    charts = ws._charts
    assert len(charts) == 1
    assert charts[0].type == "col"  # >=3 numeric cols -> column variant


def test_no_chart_for_table_only_shape():
    cols = [{"name": "A", "type": "string"}, {"name": "B", "type": "string"}]
    rows = [["x", "y"]]
    ws = _reopen(build_xlsx_report(cols, rows))["QueryResults"]
    assert len(ws._charts) == 0


def test_empty_rowset_guard_no_chart_but_valid():
    cols = [{"name": "TimeGenerated", "type": "datetime"}, {"name": "V", "type": "long"}]
    ws = _reopen(build_xlsx_report(cols, []))["QueryResults"]
    assert len(ws._charts) == 0
    assert [c.value for c in ws[1]] == ["TimeGenerated", "V"]  # header still present


# ── Non-fatal chart failure ──────────────────────────────────────────────────
def test_chart_build_failure_still_yields_valid_workbook(monkeypatch):
    def boom(*args, **kwargs):
        raise RuntimeError("simulated chart failure")

    monkeypatch.setattr(xlsx_report, "_build_and_add_chart", boom)

    cols = [{"name": "TimeGenerated", "type": "datetime"}, {"name": "V", "type": "long"}]
    rows = [["2026-08-01T00:00:00Z", 10], ["2026-08-01T01:00:00Z", 20]]
    data = build_xlsx_report(cols, rows)  # must NOT raise
    ws = _reopen(data)["QueryResults"]
    assert len(ws._charts) == 0  # chart step failed -> no chart
    # data sheet is intact and typed
    assert [c.value for c in ws[1]] == ["TimeGenerated", "V"]
    assert ws.cell(row=2, column=1).is_date is True
    assert ws.cell(row=2, column=2).value == 10
